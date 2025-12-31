# pip install pandas openpyxl python-dateutil

import os
import re
import calendar
import pandas as pd
from datetime import date, datetime
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook

# =========================
# CONFIG
# =========================
EXCEL_PATH = r"C:\usuario\95677\local.xlsx"
SHEET_TRIMESTRAL = "Trimestral"
SHEET_DATOS = "Datos"
CSV_DIR = r"C:\archivos\csv"

# =========================
# Mes siguiente (A2:A10)
# =========================
MONTHS = [m for m in calendar.month_name if m]  # January..December
MONTH_PATTERN = re.compile(r"\b(" + "|".join(MONTHS) + r")\b", flags=re.IGNORECASE)

def next_month_name(month_name: str) -> str:
    idx = {m.lower(): i for i, m in enumerate(MONTHS, start=1)}.get(month_name.lower())
    if not idx:
        return month_name
    next_idx = 1 if idx == 12 else idx + 1
    return MONTHS[next_idx - 1]

def replace_month_with_next(value):
    if value is None:
        return value

    if isinstance(value, (datetime, date)):
        return next_month_name(MONTHS[value.month - 1])

    if isinstance(value, str):
        m = MONTH_PATTERN.search(value)
        if not m:
            return value
        found = m.group(1)
        return MONTH_PATTERN.sub(next_month_name(found), value, count=1)

    return value

# =========================
# 1) Modificar "Trimestral"
# =========================
def update_trimestral(excel_path: str) -> None:
    wb = load_workbook(excel_path)
    if SHEET_TRIMESTRAL not in wb.sheetnames:
        raise ValueError(f"No existe la hoja '{SHEET_TRIMESTRAL}'. Hojas: {wb.sheetnames}")
    ws = wb[SHEET_TRIMESTRAL]

    # Para leer solo valores (resultado cacheado de fórmulas)
    wb_vals = load_workbook(excel_path, data_only=True)
    ws_vals = wb_vals[SHEET_TRIMESTRAL]

    # A2:A10 -> mes siguiente (EN)
    for r in range(2, 11):
        ws[f"A{r}"].value = replace_month_with_next(ws[f"A{r}"].value)

    # C5:E10 -> C2:E7 (SIEMPRE valores, NUNCA fórmulas)
    cols = ["C", "D", "E"]
    missing_cached = []

    for i in range(6):  # filas 5..10 -> 2..7
        src_row = 5 + i
        dst_row = 2 + i
        for col in cols:
            v = ws_vals[f"{col}{src_row}"].value  # valor calculado/cacheado
            if v is None:
                ws[f"{col}{dst_row}"].value = None
                missing_cached.append(f"{col}{src_row}")
            else:
                ws[f"{col}{dst_row}"].value = v

    wb.save(excel_path)

    if missing_cached:
        print(
            "[AVISO] Algunas celdas origen no tenían valor cacheado (data_only=True devolvió None).\n"
            "No se copió ninguna fórmula; esas celdas destino se dejaron en blanco.\n"
            f"Celdas afectadas: {', '.join(missing_cached)}\n"
            "Solución: abre el XLSX en Excel, recalcula si hace falta y guarda; luego ejecuta de nuevo."
        )

# =========================
# 2) Leer CSVs y volcar a "Datos"
# =========================
def build_df_from_csvs(csv_dir: str) -> pd.DataFrame:
    # Fecha del 1º día del mes anterior (formato español) como tu script
    fecha_mes_anterior = (date.today().replace(day=1) - relativedelta(months=1)).strftime("%d/%m/%Y")

    columnas_finales = [
        "ID_REG", "ID_ORG", "Fecha", "Organización", "ID_Cluster",
        "Total Attempted Messages", "Total Threat Messages", "Clean Messages",
        "Stopped by Reputation Filtering", "Stopped as Invalid Recipients",
        "Spam Detected", "Virus Detected", "Detected by Advanced Malware Protection",
        "Stopped by Content Filter", "Messages with Malicious URLs", "Stopped by DMARC"
    ]

    dataframes = []
    id_reg = 1

    for archivo in os.listdir(csv_dir):
        if not archivo.lower().endswith(".csv"):
            continue

        ruta = os.path.join(csv_dir, archivo)

        try:
            # Encabezado fila 1, saltar fila 2 (como tu script)
            df = pd.read_csv(ruta, header=0, skiprows=[1])

            # Asegurar que existe la fila 3 (índice 2)
            if len(df) < 3:
                empty_row = pd.Series([0] * len(df.columns), index=df.columns)
                while len(df) < 3:
                    df = pd.concat([df, pd.DataFrame([empty_row])], ignore_index=True)

            if df.iloc[2].isnull().all():
                empty_row = pd.Series([0] * len(df.columns), index=df.columns)
                df.iloc[2] = empty_row

            # Extraer Organización e ID_Cluster del nombre del archivo: ORG-CLUSTER.csv
            base = os.path.splitext(archivo)[0]
            parts = base.split("-")
            organizacion = parts[0].strip() if len(parts) >= 1 else ""
            id_cluster = parts[1].strip() if len(parts) >= 2 else ""

            fila = {
                "ID_REG": id_reg,
                "ID_ORG": id_reg,  # mantengo tu lógica
                "Fecha": fecha_mes_anterior,
                "Organización": organizacion,
                "ID_Cluster": id_cluster,

                "Total Attempted Messages": df.get("Total Attempted Messages", pd.Series([0])).iloc[0],
                "Total Threat Messages": df.get("Total Threat Messages", pd.Series([0])).iloc[0],
                "Clean Messages": df.get("Clean Messages", pd.Series([0])).iloc[0],
                "Stopped by Reputation Filtering": df.get("Stopped by Reputation Filtering", pd.Series([0])).iloc[0],
                "Stopped as Invalid Recipients": df.get("Stopped as Invalid Recipients", pd.Series([0])).iloc[0],
                "Spam Detected": df.get("Spam Detected", pd.Series([0])).iloc[0],
                "Virus Detected": df.get("Virus Detected", pd.Series([0])).iloc[0],
                "Detected by Advanced Malware Protection": df.get("Detected by Advanced Malware Protection", pd.Series([0])).iloc[0],
                "Stopped by Content Filter": df.get("Stopped by Content Filter", pd.Series([0])).iloc[0],
                "Messages with Malicious URLs": df.get("Messages with Malicious URLs", pd.Series([0])).iloc[0],
                "Stopped by DMARC": df.get("Stopped by DMARC", pd.Series([0])).iloc[0],
            }

            dataframes.append(pd.DataFrame([fila]))
            id_reg += 1

        except Exception as e:
            print(f"[WARN] Error procesando {archivo}: {e}")

    if not dataframes:
        raise RuntimeError("No se encontraron CSV válidos en la carpeta indicada.")

    return pd.concat(dataframes, ignore_index=True)[columnas_finales]

def write_df_to_excel(excel_path: str, df: pd.DataFrame) -> None:
    # Reemplaza SOLO la hoja 'Datos', mantiene el resto
    with pd.ExcelWriter(excel_path, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name=SHEET_DATOS, index=False)

# =========================
# MAIN (orden solicitado)
# =========================
def main():
    # 1) Primero Trimestral
    update_trimestral(EXCEL_PATH)

    # 2) Después Datos desde CSVs
    df = build_df_from_csvs(CSV_DIR)
    write_df_to_excel(EXCEL_PATH, df)

    print("OK: 'Trimestral' actualizado y CSVs consolidados en 'Datos'.")

if __name__ == "__main__":
    main()