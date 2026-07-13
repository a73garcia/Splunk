from pathlib import Path
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl import load_workbook

from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment
)

from openpyxl.utils import get_column_letter

from config import *
from models import Task


class ExcelManager:

    def __init__(self):

        self.file = EXCEL_FILE

        self.header_fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78"
        )

        self.header_font = Font(
            bold=True,
            color="FFFFFF"
        )

        self.border = Border(

            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")

        )

        self.center = Alignment(
            horizontal="center",
            vertical="center"
        )

        if not self.file.exists():

            self.create_database()

    # ==========================================================
    # LIBRO
    # ==========================================================

    def workbook(self):

        return load_workbook(self.file)

    def save(self, workbook):

        workbook.save(self.file)

    # ==========================================================
    # CREAR BASE DE DATOS
    # ==========================================================

    def create_database(self):

        workbook = Workbook()

        self.create_tasks_sheet(workbook)

        self.create_history_sheet(workbook)

        self.create_metrics_sheet(workbook)

        self.create_owners_sheet(workbook)

        self.create_categories_sheet(workbook)

        self.create_calendar_sheet(workbook)

        self.create_config_sheet(workbook)

        workbook.save(self.file)

    # ==========================================================
    # FORMATO
    # ==========================================================

    def format_header(self, worksheet):

        for cell in worksheet[1]:

            cell.fill = self.header_fill

            cell.font = self.header_font

            cell.border = self.border

            cell.alignment = self.center

        for column in range(1, worksheet.max_column + 1):

            worksheet.column_dimensions[
                get_column_letter(column)
            ].width = 20

    # ==========================================================
    # HOJA TAREAS
    # ==========================================================

    def create_tasks_sheet(self, workbook):

        ws = workbook.active

        ws.title = SHEET_TASKS

        ws.append([

            "ID",

            "Título",

            "Descripción",

            "Responsable",

            "Estado",

            "Prioridad",

            "Categoría",

            "Etiquetas",

            "Proyecto",

            "Fecha Creación",

            "Fecha Inicio",

            "Fecha Prevista",

            "Fecha Finalización",

            "Avance",

            "Comentarios"

        ])

        self.format_header(ws)

    # ==========================================================
    # HOJA HISTORIAL
    # ==========================================================

    def create_history_sheet(self, workbook):

        ws = workbook.create_sheet(

            SHEET_HISTORY

        )

        ws.append([

            "TaskID",

            "Fecha",

            "Usuario",

            "Avance",

            "Comentario"

        ])

        self.format_header(ws)

    # ==========================================================
    # HOJA MÉTRICAS
    # ==========================================================

    def create_metrics_sheet(self, workbook):

        ws = workbook.create_sheet(

            SHEET_METRICS

        )

        ws.append([

            "Fecha",

            "Pendientes",

            "En curso",

            "Bloqueadas",

            "Finalizadas",

            "Retrasadas",

            "Total",

            "Avance"

        ])

        self.format_header(ws)

    # ==========================================================
    # HOJA RESPONSABLES
    # ==========================================================

    def create_owners_sheet(self, workbook):

        ws = workbook.create_sheet(

            SHEET_OWNERS

        )

        ws.append([

            "Responsable",

            "Departamento",

            "Correo"

        ])

        self.format_header(ws)

    # ==========================================================
    # HOJA CATEGORÍAS
    # ==========================================================

    def create_categories_sheet(self, workbook):

        ws = workbook.create_sheet(

            SHEET_CATEGORIES

        )

        ws.append([

            "Categoría",

            "Color"

        ])

        self.format_header(ws)

    # ==========================================================
    # HOJA CALENDARIO
    # ==========================================================

    def create_calendar_sheet(self, workbook):

        ws = workbook.create_sheet(

            SHEET_CALENDAR

        )

        ws.append([

            "Fecha",

            "Año",

            "Mes",

            "NombreMes",

            "Semana",

            "Día",

            "NombreDia"

        ])

        self.format_header(ws)

        fecha = datetime(2025, 1, 1)

        while fecha.year <= 2035:

            ws.append([

                fecha.strftime(DATE_FORMAT),

                fecha.year,

                fecha.month,

                fecha.strftime("%B"),

                fecha.isocalendar()[1],

                fecha.day,

                fecha.strftime("%A")

            ])

            fecha += timedelta(days=1)

    # ==========================================================
    # HOJA CONFIGURACIÓN
    # ==========================================================

    def create_config_sheet(self, workbook):

        ws = workbook.create_sheet(

            SHEET_CONFIG

        )

        ws["A1"] = APP_NAME

        ws["A2"] = "Versión"

        ws["B2"] = APP_VERSION

        ws["A3"] = "Creado"

        ws["B3"] = datetime.now().strftime(

            DATETIME_FORMAT

        )

        self.format_header(ws)
        
    # ==========================================================
    # SIGUIENTE ID
    # ==========================================================

    def next_id(self):

        tareas = self.load_tasks()

        if not tareas:
            return 1

        return max(t.id for t in tareas) + 1

    # ==========================================================
    # CARGAR TAREAS
    # ==========================================================

    def load_tasks(self):

        workbook = self.workbook()

        ws = workbook[SHEET_TASKS]

        tareas = []

        for row in ws.iter_rows(min_row=2, values_only=True):

            if row[0] is None:
                continue

            tarea = Task()

            tarea.id = row[0]
            tarea.titulo = row[1] or ""
            tarea.descripcion = row[2] or ""
            tarea.responsable = row[3] or ""
            tarea.estado = row[4] or STATUS_PENDING
            tarea.prioridad = row[5] or PRIORITY_MEDIUM
            tarea.categoria = row[6] or ""
            tarea.etiquetas = row[7] or ""
            tarea.fecha_creacion = row[9] or ""
            tarea.fecha_inicio = row[10] or ""
            tarea.fecha_prevista = row[11] or ""
            tarea.fecha_finalizacion = row[12] or ""
            tarea.avance = row[13] or 0
            tarea.comentarios = row[14] or ""

            tareas.append(tarea)

        workbook.close()

        return tareas

    # ==========================================================
    # BUSCAR TAREA
    # ==========================================================

    def get_task(self, task_id):

        for tarea in self.load_tasks():

            if tarea.id == task_id:
                return tarea

        return None

    # ==========================================================
    # INSERTAR TAREA
    # ==========================================================

    def add_task(self, task):

        workbook = self.workbook()

        ws = workbook[SHEET_TASKS]

        if task.id == 0:
            task.id = self.next_id()

        ws.append([

            task.id,

            task.titulo,

            task.descripcion,

            task.responsable,

            task.estado,

            task.prioridad,

            task.categoria,

            task.etiquetas,

            "",

            task.fecha_creacion,

            task.fecha_inicio,

            task.fecha_prevista,

            task.fecha_finalizacion,

            task.avance,

            task.comentarios

        ])

        self.save(workbook)

    # ==========================================================
    # ACTUALIZAR TAREA
    # ==========================================================

    def update_task(self, task):

        workbook = self.workbook()

        ws = workbook[SHEET_TASKS]

        for row in range(2, ws.max_row + 1):

            if ws.cell(row, 1).value != task.id:
                continue

            ws.cell(row, 2).value = task.titulo
            ws.cell(row, 3).value = task.descripcion
            ws.cell(row, 4).value = task.responsable
            ws.cell(row, 5).value = task.estado
            ws.cell(row, 6).value = task.prioridad
            ws.cell(row, 7).value = task.categoria
            ws.cell(row, 8).value = task.etiquetas
            ws.cell(row,10).value = task.fecha_creacion
            ws.cell(row,11).value = task.fecha_inicio
            ws.cell(row,12).value = task.fecha_prevista
            ws.cell(row,13).value = task.fecha_finalizacion
            ws.cell(row,14).value = task.avance
            ws.cell(row,15).value = task.comentarios

            break

        self.save(workbook)

    # ==========================================================
    # ELIMINAR TAREA
    # ==========================================================

    def delete_task(self, task_id):

        workbook = self.workbook()

        ws = workbook[SHEET_TASKS]

        for row in range(2, ws.max_row + 1):

            if ws.cell(row, 1).value == task_id:

                ws.delete_rows(row)

                break

        self.save(workbook)

    # ==========================================================
    # EXISTE TAREA
    # ==========================================================

    def task_exists(self, task_id):

        return self.get_task(task_id) is not None

    # ==========================================================
    # TOTAL TAREAS
    # ==========================================================

    def task_count(self):

        return len(self.load_tasks())

    # ==========================================================
    # TAREAS POR ESTADO
    # ==========================================================

    def tasks_by_status(self, estado):

        return [

            t

            for t in self.load_tasks()

            if t.estado == estado

        ]

    # ==========================================================
    # TAREAS POR RESPONSABLE
    # ==========================================================

    def tasks_by_owner(self, responsable):

        return [

            t

            for t in self.load_tasks()

            if t.responsable == responsable

        ]

    # ==========================================================
    # TAREAS RETRASADAS
    # ==========================================================

    def delayed_tasks(self):

        return [

            t

            for t in self.load_tasks()

            if t.esta_retrasada()

        ]
    
    
    # ==========================================================
    # AÑADIR HISTORIAL
    # ==========================================================

    def add_history(
        self,
        task_id,
        usuario,
        avance,
        comentario
    ):

        workbook = self.workbook()

        ws = workbook[SHEET_HISTORY]

        ws.append([

            task_id,

            datetime.now().strftime(
                DATETIME_FORMAT
            ),

            usuario,

            avance,

            comentario

        ])

        self.save(workbook)

    # ==========================================================
    # CARGAR HISTORIAL
    # ==========================================================

    def load_history(self, task_id=None):

        workbook = self.workbook()

        ws = workbook[SHEET_HISTORY]

        datos = []

        for row in ws.iter_rows(
            min_row=2,
            values_only=True
        ):

            if row[0] is None:
                continue

            if task_id is not None:

                if row[0] != task_id:
                    continue

            datos.append({

                "task_id": row[0],

                "fecha": row[1],

                "usuario": row[2],

                "avance": row[3],

                "comentario": row[4]

            })

        workbook.close()

        return datos

    # ==========================================================
    # ACTUALIZAR RESPONSABLES
    # ==========================================================

    def refresh_owners(self):

        workbook = self.workbook()

        ws = workbook[SHEET_OWNERS]

        while ws.max_row > 1:
            ws.delete_rows(2)

        responsables = sorted({

            t.responsable.strip()

            for t in self.load_tasks()

            if t.responsable.strip()

        })

        for responsable in responsables:

            ws.append([

                responsable,

                "",

                ""

            ])

        self.save(workbook)

    # ==========================================================
    # ACTUALIZAR CATEGORÍAS
    # ==========================================================

    def refresh_categories(self):

        workbook = self.workbook()

        ws = workbook[SHEET_CATEGORIES]

        while ws.max_row > 1:
            ws.delete_rows(2)

        categorias = sorted({

            t.categoria.strip()

            for t in self.load_tasks()

            if t.categoria.strip()

        })

        for categoria in categorias:

            ws.append([

                categoria,

                "#5B9BD5"

            ])

        self.save(workbook)

    # ==========================================================
    # MÉTRICAS
    # ==========================================================

    def update_metrics(self):

        workbook = self.workbook()

        ws = workbook[SHEET_METRICS]

        tareas = self.load_tasks()

        pendientes = len(

            self.tasks_by_status(
                STATUS_PENDING
            )

        )

        curso = len(

            self.tasks_by_status(
                STATUS_PROGRESS
            )

        )

        bloqueadas = len(

            self.tasks_by_status(
                STATUS_BLOCKED
            )

        )

        finalizadas = len(

            self.tasks_by_status(
                STATUS_DONE
            )

        )

        retrasadas = len(

            self.delayed_tasks()

        )

        total = len(tareas)

        if total:

            avance = round(

                sum(
                    t.avance
                    for t in tareas
                ) / total

            )

        else:

            avance = 0

        ws.append([

            datetime.now().strftime(
                DATE_FORMAT
            ),

            pendientes,

            curso,

            bloqueadas,

            finalizadas,

            retrasadas,

            total,

            avance

        ])

        self.save(workbook)

    # ==========================================================
    # REFRESCO GENERAL
    # ==========================================================

    def refresh(self):

        self.refresh_owners()

        self.refresh_categories()

        self.update_metrics()

    # ==========================================================
    # EXPORTAR SNAPSHOT
    # ==========================================================

    def snapshot(self):

        return {

            "fecha": datetime.now().strftime(
                DATETIME_FORMAT
            ),

            "total": self.task_count(),

            "pendientes": len(
                self.tasks_by_status(
                    STATUS_PENDING
                )
            ),

            "curso": len(
                self.tasks_by_status(
                    STATUS_PROGRESS
                )
            ),

            "bloqueadas": len(
                self.tasks_by_status(
                    STATUS_BLOCKED
                )
            ),

            "finalizadas": len(
                self.tasks_by_status(
                    STATUS_DONE
                )
            ),

            "retrasadas": len(
                self.delayed_tasks()
            )

        }
        
        
        
        
        
        
        
        
        
        
        
        
        
        
    # ==========================================================
    # BUSCAR TAREAS
    # ==========================================================

    def search_tasks(
        self,
        text=""
    ):

        text = text.lower().strip()

        resultado = []

        for tarea in self.load_tasks():

            contenido = " ".join([

                tarea.titulo,

                tarea.descripcion,

                tarea.responsable,

                tarea.estado,

                tarea.prioridad,

                tarea.categoria,

                tarea.etiquetas,

                tarea.comentarios

            ]).lower()

            if text in contenido:

                resultado.append(tarea)

        return resultado

    # ==========================================================
    # FILTRAR CATEGORÍA
    # ==========================================================

    def tasks_by_category(
        self,
        categoria
    ):

        return [

            t

            for t in self.load_tasks()

            if t.categoria == categoria

        ]

    # ==========================================================
    # FILTRAR PRIORIDAD
    # ==========================================================

    def tasks_by_priority(
        self,
        prioridad
    ):

        return [

            t

            for t in self.load_tasks()

            if t.prioridad == prioridad

        ]

    # ==========================================================
    # TAREAS DEL DÍA
    # ==========================================================

    def tasks_today(self):

        hoy = datetime.now().strftime(
            DATE_FORMAT
        )

        return [

            t

            for t in self.load_tasks()

            if t.fecha_prevista == hoy

        ]

    # ==========================================================
    # VALIDAR EXCEL
    # ==========================================================

    def validate_database(self):

        workbook = self.workbook()

        required = [

            SHEET_TASKS,

            SHEET_HISTORY,

            SHEET_METRICS,

            SHEET_OWNERS,

            SHEET_CATEGORIES,

            SHEET_CALENDAR,

            SHEET_CONFIG

        ]

        for hoja in required:

            if hoja not in workbook.sheetnames:

                workbook.close()

                return False

        workbook.close()

        return True

    # ==========================================================
    # REPARAR BASE DE DATOS
    # ==========================================================

    def repair_database(self):

        if self.validate_database():

            return

        self.create_database()

    # ==========================================================
    # BACKUP
    # ==========================================================

    def create_backup(self):

        import shutil

        nombre = datetime.now().strftime(
            "Backup_%Y%m%d_%H%M%S.xlsx"
        )

        destino = BACKUP_DIR / nombre

        shutil.copy2(

            self.file,

            destino

        )

        return destino

    # ==========================================================
    # ESTADÍSTICAS
    # ==========================================================

    def statistics(self):

        tareas = self.load_tasks()

        return {

            "total": len(tareas),

            "pendientes": len(
                self.tasks_by_status(
                    STATUS_PENDING
                )
            ),

            "curso": len(
                self.tasks_by_status(
                    STATUS_PROGRESS
                )
            ),

            "bloqueadas": len(
                self.tasks_by_status(
                    STATUS_BLOCKED
                )
            ),

            "finalizadas": len(
                self.tasks_by_status(
                    STATUS_DONE
                )
            ),

            "retrasadas": len(
                self.delayed_tasks()
            ),

            "responsables": len({

                t.responsable

                for t in tareas

            }),

            "categorias": len({

                t.categoria

                for t in tareas

            })

        }

    # ==========================================================
    # INFORMACIÓN
    # ==========================================================

    def database_info(self):

        return {

            "archivo": str(self.file),

            "hojas": self.workbook().sheetnames,

            "registros": self.task_count(),

            "ultima_actualizacion":

                datetime.now().strftime(

                    DATETIME_FORMAT

                )

        }

    # ==========================================================
    # MANTENIMIENTO
    # ==========================================================

    def maintenance(self):

        self.repair_database()

        self.refresh()

        self.create_backup()

    # ==========================================================
    # CIERRE
    # ==========================================================

    def close(self):

        pass